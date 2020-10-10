from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.uic import loadUiType
from os import path
import sys
import threading
import xlsxwriter
import xlrd
from openpyxl import *
import time
import pymysql

Form_Class,_ = loadUiType(path.join(path.dirname(__file__), "main.ui"))

class mainapp(QMainWindow, Form_Class):

	def __init__(self, parent=None):
		super(mainapp,self).__init__(parent)
		QMainWindow.__init__(self)
		self.setupUi(self)
		self.handle_UI()
		#self.menuHello.setTitle("ahmed")
		self.buttons()
		self.thread = Threads()


		# Salary digits Validation

	def handle_UI(self):
		self.setWindowTitle('AH Employee Regestration v2.0')
		self.setFixedSize(577,628)

	def buttons(self):
		self.pushButton.clicked.connect(self.check)
		self.pushButton_2.clicked.connect(self.search)
		self.pushButton_3.clicked.connect(self.reset)


	def txtmaker(self):
		self.excel_db()
		svid = self.lineEdit.text()
		svname = self.lineEdit_2.text()
		svaddress = self.lineEdit_3.text()
		svSalary = self.lineEdit_4.text()
		svage = self.lineEdit_6.text()
		file_name = svid + '_' + svname + '.txt'
		f = open('data/'+file_name, 'w')
		f.write('id   : ' + svid + '\n')
		f.write('Name   : ' + svname + '\n')
		f.write('Address: ' + svaddress + '\n')
		f.write('Salary : ' + svSalary + '\n')
		f.write('Age    : ' + svage + '\n')
		f.close()



	def xlsxmaker(self):

		self.excel_db()
		svid = self.lineEdit.text()
		svname = self.lineEdit_2.text()
		svaddress = self.lineEdit_3.text()
		svSalary = self.lineEdit_4.text()
		svage = self.lineEdit_6.text()
		file_name = svid + '_' + svname + '.xlsx'
		f = xlsxwriter.Workbook('data/'+file_name)
		out = f.add_worksheet()
		out.write('A1','id   : ')
		out.write('A2','Name   : ')
		out.write('A3','Address: ')
		out.write('A4','Salary : ')
		out.write('A5','Age : ')

		out.write('B1',svid)
		out.write('B2',svname)
		out.write('B3',svaddress)
		out.write('B4',svSalary)
		out.write('B5',svage)
		f.close()
	
	def employee_db(self):
		svid = str(self.lineEdit.text())
		svname = str(self.lineEdit_2.text())
		svaddress =str(self.lineEdit_3.text())
		svSalary = self.lineEdit_4.text()
		svage = self.lineEdit_6.text()
		conn = pymysql.connect(host="localhost",user="root",passwd="",database="py_db")
		cur = conn.cursor()
		query = ("insert into employees values(%s,%s,%s,%s,%s)")
		data = cur.execute(query,(svid,svname,svaddress,svSalary,svage))
		conn.commit()
		if (data):
			print("Successfully")

	def search(self):
		try:
			conn = mysql.connector.connect(user='root', passwd='', database='py_db')
			cur = conn.cursor()
			cur.execute("SELECT * FROM employees")
			for x in cur:
				emps = str(x[0])
				print(x)
				print(type(x))

				find = self.lineEdit_5.text()
				if find == x[0]:
					print("okkk")

		except Exception:
			QMessageBox.warning(self,'incorrect id','Employee not Found!')



	def excel_db(self):

		svid = self.lineEdit.text()
		svname = self.lineEdit_2.text()
		svaddress = self.lineEdit_3.text()
		svSalary = self.lineEdit_4.text()
		svage = self.lineEdit_6.text()

		if path.exists("db/db.xlsx"):
			
			wb = load_workbook("db/db.xlsx")
			sheet = wb.active

		else:
			wb = Workbook()
			sheet = wb.active

			data = [("id","name","address","salary","age")]
			sheet.append(data[0])

		for r in range(1,sheet.max_row+1):
			ids = sheet.cell(row=r,column=1).value

		if self.checkBox.isChecked() and self.checkBox_2.isChecked():
			r = sheet.max_row + 1

			if ids != "id":
				r = sheet.max_row

			sheet['A'+ str(r)] = svid
			sheet['B'+ str(r)] = svname
			sheet['C'+ str(r)] = svaddress
			sheet['D'+ str(r)] = svSalary
			sheet['E'+ str(r)] = svage

		elif self.checkBox.isChecked():

			r = sheet.max_row + 1

			if svid == ids:
				p = "Do you want to update %s's Data" % svname
				QMessageBox.warning(self,'Employee Found!',p)
				r = sheet.max_row

			sheet['A'+ str(r)] = svid
			sheet['B'+ str(r)] = svname
			sheet['C'+ str(r)] = svaddress
			sheet['D'+ str(r)] = svSalary
			sheet['E'+ str(r)] = svage

		elif self.checkBox_2.isChecked():
			if svid == ids:
				p = "Do you want to update %s's Data" % svname
				QMessageBox.warning(self,'Employee Found!',p)

			r = sheet.max_row + 1
			sheet['A'+ str(r)] = svid
			sheet['B'+ str(r)] = svname
			sheet['C'+ str(r)] = svaddress
			sheet['D'+ str(r)] = svSalary
			sheet['E'+ str(r)] = svage

		

		wb.save('db/db.xlsx')


	def check(self):

		#self.thread.start()
		#self.employee_db()

		svid = self.lineEdit.text()
		svname = self.lineEdit_2.text()
		svaddress = self.lineEdit_3.text()
		svSalary = self.lineEdit_4.text()
		svage = self.lineEdit_6.text()

		file_name1 = svid + '_' + svname + '.txt'
		file_name2 = svid + '_' + svname + '.xlsx'

		if self.lineEdit.text().strip()== '':
			QMessageBox.warning(self,'Enter id','id is Empty !')
			self.lineEdit.setFocus()
		elif self.lineEdit_2.text().strip() == '':
			QMessageBox.warning(self,'Enter name','name field is Empty !')
			self.lineEdit_2.setFocus()
		elif self.lineEdit_3.text().strip() == '':
			QMessageBox.warning(self,'Enter Address','Address field is Empty !')
			self.lineEdit_3.setFocus()
		elif self.lineEdit_4.text().strip() == '':
			QMessageBox.warning(self,'Enter Salary','Salary field is Empty !')
			self.lineEdit_4.setFocus()

		elif svSalary.isdecimal() == False:
			QMessageBox.warning(self,'Error numeric value','Enter numeric value in Salary')
			self.lineEdit_4.setFocus()

		elif self.lineEdit_6.text().strip() == '':
			QMessageBox.warning(self,'Enter Age','Age field is Empty !')
			self.lineEdit_6.setFocus()

		elif svage.isdecimal() == False:
			QMessageBox.warning(self,'Error numeric value','Enter numeric value in Age')
			self.lineEdit_6.setFocus()

		else:


			if self.checkBox.isChecked():

				if path.exists('data/'+ file_name1):

					QMessageBox.warning(self,'File already exists','....')
					self.txtmaker()
					QMessageBox.information(self,'Done !','Employee Text File updated Successfully !')
					self.lineEdit.setFocus()
				else:
					self.txtmaker()
					QMessageBox.information(self,'Done !','Employee Text File Created Successfully !')
					self.lineEdit.setFocus()


			if self.checkBox_2.isChecked():

				if path.exists('data/'+ file_name2):
					QMessageBox.warning(self,'File already exists','....')
					self.xlsxmaker()
					QMessageBox.information(self,'Done !','Employee Excel File updated Successfully !')
					self.lineEdit.setFocus()
				else:
					self.xlsxmaker()
					QMessageBox.information(self,'Done !','Employee Excel File Created Successfully !')
					self.lineEdit.setFocus()

			elif self.checkBox.isChecked() == False:
				if self.checkBox_2.isChecked() == False:
					QMessageBox.warning(self,'choose error','Choose the Export file type. i.e you can choose them together')

				
	def reset(self):
		self.lineEdit.setText('')
		self.lineEdit_2.setText('')
		self.lineEdit_3.setText('')
		self.lineEdit_4.setText('')
		self.lineEdit.setFocus()



Form_Class2,_ = loadUiType(path.join(path.dirname(__file__), "main2.ui"))

class signin(QMainWindow, Form_Class2):

	def __init__(self, parent=None):
		super(signin,self).__init__(parent)
		QMainWindow.__init__(self)
		self.setupUi(self)
		self.handle_UI()
		self.buttons()

	def handle_UI(self):
		self.setWindowTitle('Login')
		self.setFixedSize(461,461)

	def buttons(self):
		self.pushButton.clicked.connect(self.login)

	def login(self):
		if self.lineEdit.text() == "admin":
			if self.lineEdit_2.text() == "pass":
				self.index = mainapp()
				self.hide()
				self.index.show()



class Threads(QThread):
	def __init__(self,parent=None):
		super(Threads, self).__init__(parent)

	def run(self):
		print("threaded")

def main():

	app = QApplication(sys.argv)
	window = signin()
	window.show()
	sys.exit(app.exec_())

if __name__ == "__main__":
	main()

